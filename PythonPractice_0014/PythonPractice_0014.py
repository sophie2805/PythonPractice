#! /usr/bin/env python
#! -*- coding: utf-8 -*-

__author__ = 'Sophie'

from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,colors
import re

wb = Workbook()

dest_file = '/Users/Sophie/PycharmProjects/PythonPractice/PythonPractice_0014/student.xlsx'

ws = wb.create_sheet(title = 'student')

alignment = Alignment(horizontal='center', vertical='center')
font = Font(bold=True,color=colors.BLUE)

col = 'ABCDE'

ws[col[0]+'1'] = '#'
ws[col[1]+'1'] = '姓名'
ws[col[2]+'1'] = '数学'
ws[col[3]+'1'] = '英语'
ws[col[4]+'1'] = '语文'

for x in col:
    ws[x+'1'].alignment = alignment
    ws[x+'1'].font = font

data = []

p = re.compile(':\[')

source_file = open('/Users/Sophie/PycharmProjects/PythonPractice/PythonPractice_0014/student.txt','r')

for line in source_file:
    line = line.strip('     \n')
    if len(line) > 0 and not line.startswith('{') and not line.startswith('}'):
        line = line.strip('], ')
        line = p.sub(',',line)
        data.append(line.split(','))

#print data
source_file.close()

for i in range(len(data)):
    for j in range(5):
        ws[col[j]+str(i+2)] = data[i][j].strip('"')
        ws[col[j]+str(i+2)].alignment = alignment

wb.save(dest_file)